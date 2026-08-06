"""
WealthOS — per-AMC disclosure adapters.

An adapter per AMC is not over-engineering; it is what the files require.
Measured across five AMCs' real disclosures:

  AMC     header row   ISIN col   pct column       pct units
  Nippon      4            B      "% to NAV"       fraction
  SBI         6            D      "% to AUM"       PERCENTAGE
  Mirae       8            C      "% to Net Assets" fraction
  ICICI       4            C      "% to Nav"       fraction
  HDFC        5            B      "% to NAV"       PERCENTAGE

No two agree on header row or ISIN column, and they split on units with no
declaration anywhere. `pct_is_fraction` is therefore explicit per adapter and
must never be inferred from the header caption.

Packaging differs too: one 108-sheet workbook, one 124-sheet workbook, one file
per scheme, and a ZIP of legacy .xls.

methodology_version: amfi_ingest@1.0.0
"""

from __future__ import annotations

import io
import re
import logging
from datetime import date, datetime
from typing import Iterable, Optional

import openpyxl

from app.crawler.core import ISIN_RE, ParsedSheet, Row

logger = logging.getLogger(__name__)

_DATE_PATTERNS = [
    (re.compile(r"(\d{1,2})[-/\s]([A-Za-z]{3,9})[-/\s,]*(\d{4})"), "dmy"),
    (re.compile(r"([A-Za-z]{3,9})\s+(\d{1,2})\s*,\s*(\d{4})"), "mdy"),
    (re.compile(r"(\d{4})-(\d{2})-(\d{2})"), "iso"),
]
_MONTHS = {m.lower(): i for i, m in enumerate(
    ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"], 1)}


def parse_date(v) -> Optional[date]:
    """Accept a real datetime cell or the many free-text forms AMCs use."""
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v or "")
    for rx, kind in _DATE_PATTERNS:
        m = rx.search(s)
        if not m:
            continue
        try:
            if kind == "iso":
                return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            if kind == "dmy":
                mon = _MONTHS.get(m.group(2)[:3].lower())
                if mon:
                    return date(int(m.group(3)), mon, int(m.group(1)))
            if kind == "mdy":
                mon = _MONTHS.get(m.group(1)[:3].lower())
                if mon:
                    return date(int(m.group(3)), mon, int(m.group(2)))
        except ValueError:
            continue
    return None


def load_workbook_bytes(path: str):
    """
    Load a workbook from bytes, never a path.

    openpyxl's _validate_archive rejects on FILE EXTENSION, and extensions lie:
    Nippon ships XLSX named `.xls`, SBI alternates `.xls`/`.xlsx` between
    months, DSP serves a ZIP from a `.xlsx` link. Loading via BytesIO bypasses
    the extension check so content decides.
    """
    with open(path, "rb") as f:
        raw = f.read()
    if raw[:4] != b"PK\x03\x04":
        raise ValueError(f"not an OOXML workbook (magic={raw[:4]!r})")
    return openpyxl.load_workbook(io.BytesIO(raw), data_only=True)


class BaseAdapter:
    """
    Extract per-scheme constituent rows from one AMC's disclosure file.

    Adapters never write to the database and never decide what is valid —
    they parse and hand ParsedSheet to the gates.
    """

    amc_code: str = "unknown"
    header_row: int = 1
    col_isin: int = 1
    col_name: int = 2
    col_industry: Optional[int] = None
    col_qty: Optional[int] = None
    col_value: Optional[int] = None
    col_pct: int = 3
    pct_is_fraction: bool = True     # EXPLICIT. Never inferred.
    data_start_offset: int = 1       # rows after header where data may begin

    # ── hooks ──
    def sheet_names(self, wb) -> Iterable[str]:
        return wb.sheetnames

    def scheme_name(self, ws, sheet: str) -> str:
        raise NotImplementedError

    def disclosure_date(self, ws) -> Optional[date]:
        raise NotImplementedError

    # ── shared extraction ──
    def _rows(self, ws) -> list[Row]:
        out: list[Row] = []
        for r in range(self.header_row + self.data_start_offset, ws.max_row + 1):
            isin = ws.cell(row=r, column=self.col_isin).value
            if not isin or not isinstance(isin, str):
                continue
            isin = isin.strip()
            if not ISIN_RE.match(isin):
                continue
            pct = ws.cell(row=r, column=self.col_pct).value
            if not isinstance(pct, (int, float)):
                continue                      # 'NIL' and blanks are not data
            pct = float(pct) * (100.0 if self.pct_is_fraction else 1.0)
            out.append(Row(
                isin=isin,
                name=str(ws.cell(row=r, column=self.col_name).value or "").strip(),
                industry=(str(ws.cell(row=r, column=self.col_industry).value).strip()
                          if self.col_industry and ws.cell(row=r, column=self.col_industry).value else None),
                quantity=_num(ws.cell(row=r, column=self.col_qty).value) if self.col_qty else None,
                value=_num(ws.cell(row=r, column=self.col_value).value) if self.col_value else None,
                pct=pct,
            ))
        return _aggregate(out)

    def parse_file(self, path: str) -> list[ParsedSheet]:
        wb = load_workbook_bytes(path)
        sheets: list[ParsedSheet] = []
        for sn in self.sheet_names(wb):
            try:
                ws = wb[sn]
                rows = self._rows(ws)
                if not rows:
                    continue
                sheets.append(ParsedSheet(
                    amc_code=self.amc_code,
                    scheme_name=self.scheme_name(ws, sn),
                    disclosure_date=self.disclosure_date(ws),
                    rows=rows,
                    source_file=path,
                    source_sheet=sn,
                ))
            except Exception as e:
                logger.warning("%s: sheet %r failed: %s", self.amc_code, sn, e)
        return sheets


def _num(v):
    return float(v) if isinstance(v, (int, float)) else None


def _aggregate(rows: list[Row]) -> list[Row]:
    """
    Collapse repeated ISINs within one sheet, summing weight, quantity and value.

    The same security legitimately occupies several rows — locked-in vs free
    shares being the common case. Summing is the only correct reading: they are
    one position in one fund. Order is preserved by first appearance so output
    stays deterministic (CLAUDE.md 2.1).
    """
    merged: dict[str, Row] = {}
    for r in rows:
        cur = merged.get(r.isin)
        if cur is None:
            merged[r.isin] = r
            continue
        cur.pct += r.pct
        if r.quantity is not None:
            cur.quantity = (cur.quantity or 0) + r.quantity
        if r.value is not None:
            cur.value = (cur.value or 0) + r.value
        cur.industry = cur.industry or r.industry
    return list(merged.values())


class ICICIAdapter(BaseAdapter):
    """
    One file per scheme. Verified on `ICICI Prudential Large & Mid Cap Fund.xlsx`.

      r2  scheme name
      r3  'Portfolio as on Jun 30,2026'
      r4  HEADER: Name | ISIN | Coupon | Industry/Rating | Quantity | Exposure | % to Nav
    """
    amc_code = "icici"
    header_row = 4
    col_name, col_isin, col_industry, col_qty, col_value, col_pct = 2, 3, 5, 6, 7, 8
    pct_is_fraction = True                      # measured: sum ≈ 0.9509

    def sheet_names(self, wb):
        # 'Derivative' sheets repeat exposure and would double-count.
        return [s for s in wb.sheetnames if "derivativ" not in s.lower()]

    def scheme_name(self, ws, sheet):
        return str(ws.cell(row=2, column=2).value or sheet).strip()

    def disclosure_date(self, ws):
        return parse_date(ws.cell(row=3, column=2).value)


class HDFCAdapter(BaseAdapter):
    """
    One file per scheme. Verified on `Monthly HDFC Mid Cap Fund - 30 June 2026.xlsx`.

      r1  scheme name
      r2  'Portfolio as on 30-Jun-2026'
      r5  HEADER: ISIN | Coupon | Name | Industry+/Rating | Quantity | Value | % to NAV
    """
    amc_code = "hdfc"
    header_row = 5
    col_isin, col_name, col_industry, col_qty, col_value, col_pct = 2, 4, 5, 6, 7, 8
    pct_is_fraction = False                     # measured: sum ≈ 92.06

    def scheme_name(self, ws, sheet):
        return str(ws.cell(row=1, column=1).value or sheet).strip()

    def disclosure_date(self, ws):
        return parse_date(ws.cell(row=2, column=1).value)


class NipponAdapter(BaseAdapter):
    """
    One workbook, ~108 sheets, sheet 1 is an Index. Verified on
    `NIMF-MONTHLY-PORTFOLIO-30-Jun-26.xls` (which is really XLSX).

      r1  scheme code | scheme name
      r2  'Monthly Portfolio Statement as on June 30,2026'
      r4  HEADER: ISIN | Name | Industry/Rating | Quantity | Value | % to NAV
    """
    amc_code = "nippon"
    header_row = 4
    col_isin, col_name, col_industry, col_qty, col_value, col_pct = 2, 3, 4, 5, 6, 7
    pct_is_fraction = True                      # measured: sum ≈ 0.9869

    def sheet_names(self, wb):
        return [s for s in wb.sheetnames if s.lower() != "index"]

    def scheme_name(self, ws, sheet):
        return str(ws.cell(row=1, column=2).value or sheet).strip()

    def disclosure_date(self, ws):
        return parse_date(ws.cell(row=2, column=2).value)


class SBIAdapter(BaseAdapter):
    """
    One workbook, ~124 sheets, sheet 1 is an Index. Verified on
    `all-schemes-monthly-portfolio---as-on-30th-april-2026.xlsx`.

      r3  'SCHEME NAME :' | name
      r4  'PORTFOLIO STATEMENT AS ON :' | datetime
      r6  HEADER: Name | ISIN | Rating/Industry | Quantity | Value | % to AUM
    """
    amc_code = "sbi"
    header_row = 6
    col_name, col_isin, col_industry, col_qty, col_value, col_pct = 3, 4, 5, 6, 7, 8
    pct_is_fraction = False                     # measured: sum ≈ 96.17

    def sheet_names(self, wb):
        return [s for s in wb.sheetnames if s.lower() != "index"]

    def scheme_name(self, ws, sheet):
        return str(ws.cell(row=3, column=4).value or sheet).strip()

    def disclosure_date(self, ws):
        return parse_date(ws.cell(row=4, column=4).value)


class MiraeAdapter(BaseAdapter):
    """
    One file per scheme. Verified on `mamcf-june2026.xlsx`.

      r5  scheme name | code | datetime
      r8  HEADER: Name | ISIN | Industry^/Rating | Quantity | Value | % to Net Assets
    """
    amc_code = "mirae"
    header_row = 8
    col_name, col_isin, col_industry, col_qty, col_value, col_pct = 2, 3, 4, 5, 6, 7
    pct_is_fraction = True                      # measured: sum ≈ 0.9863

    def scheme_name(self, ws, sheet):
        return str(ws.cell(row=1, column=2).value or sheet).strip()

    def disclosure_date(self, ws):
        return parse_date(ws.cell(row=5, column=6).value) or parse_date(ws.cell(row=7, column=2).value)


ADAPTERS: dict[str, BaseAdapter] = {
    a.amc_code: a() for a in (ICICIAdapter, HDFCAdapter, NipponAdapter, SBIAdapter, MiraeAdapter)
}
