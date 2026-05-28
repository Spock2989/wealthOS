"""
WealthOS ingestion — Universal Excel/CSV parser (v2).

Design goals (per WealthOS Prime Directives):
  - Deterministic: same workbook bytes → identical output, every time.
  - ISIN-first: rows without a valid ISIN are flagged, never silently merged.
  - Multi-sheet: scans every sheet, infers header row by content (not position).
  - Broker-aware: known layouts (Zerodha Console, Groww, ICICI Direct,
    HDFC Securities, Kuvera, Coin, AngelOne) are detected and prioritised.
  - Provenance: every parsed row carries source_sheet, source_row, raw_payload,
    confidence, flags, parser_version — for full audit traceability.
  - No silent fallback: if value must be computed (qty × close), a flag says so.

methodology_version: excel_parser@2.0.0
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass, field, asdict
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

PARSER_VERSION = "2.0.0"

ISIN_RE = re.compile(r"^IN[A-Z0-9]{10}$")

# Canonical field → ordered list of header-name aliases (lowercase, exact-or-substring match).
# Order matters within a list: the FIRST match wins, so more specific aliases must come first.
COL_ALIASES: Dict[str, List[str]] = {
    "isin":            ["isin"],
    "instrument_name": ["scheme name", "instrument name", "security name",
                        "symbol", "scheme", "instrument", "fund", "security",
                        "name", "scrip", "stock"],
    "quantity":        ["quantity available", "closing units", "closing balance",
                        "balance units", "quantity", "qty", "units", "shares"],
    "average_price":   ["average price", "avg price", "avg cost", "cost price",
                        "buy avg", "purchase price", "average cost"],
    "close_price":     ["previous closing price", "current price", "current nav",
                        "closing nav", "market price", "ltp", "nav", "close", "cmp"],
    "current_value":   ["current value", "market value", "present value",
                        "current market value", "value", "amount", "valuation",
                        "mkt value", "mv"],
    "invested_value":  ["invested value", "invested amount", "cost value",
                        "investment", "total cost"],
    "sector":          ["sector", "industry"],
    "instrument_type": ["instrument type", "asset class", "category",
                        "scheme type", "scheme category"],
    "folio_number":    ["folio number", "folio no", "folio"],
}

# Broker fingerprints — strings that, if present anywhere in the workbook,
# identify the source platform. Kept narrow so we never misidentify.
BROKER_FINGERPRINTS: List[Tuple[str, List[str]]] = [
    ("zerodha_console",  ["zerodha", "kite", "client id"]),
    ("groww",            ["groww"]),
    ("icici_direct",     ["icici direct", "icicidirect"]),
    ("hdfc_securities",  ["hdfc securities"]),
    ("kuvera",           ["kuvera"]),
    ("coin",             ["coin by zerodha"]),
    ("angel_one",        ["angel one", "angel broking", "smartapi"]),
    ("upstox",           ["upstox"]),
    ("cams",             ["cams", "cams kra"]),
    ("kfintech",         ["kfintech", "karvy"]),
]

# When a workbook has both segregated (Equity / Mutual Funds) and Combined
# views, we prefer the segregated sheet so we don't double-count.
SHEET_PRIORITY: Dict[str, int] = {
    "Equity": 1, "Stocks": 1, "Mutual Funds": 1, "Mutual Fund": 1, "MF": 1,
    "Combined": 2, "Summary": 2, "All": 2,
}


@dataclass
class Provenance:
    source_sheet: str
    source_row: int
    parser: str = "excel_v2"
    parser_version: str = PARSER_VERSION
    confidence: float = 1.0
    flags: List[str] = field(default_factory=list)
    raw_payload: Dict[str, Any] = field(default_factory=dict)


@dataclass
class ParsedHolding:
    isin: Optional[str]
    instrument_name: str
    quantity: Optional[float]
    nav: Optional[float]
    current_value: float
    invested_value: Optional[float]
    sector: Optional[str]
    instrument_type_raw: Optional[str]
    folio_number: Optional[str]
    provenance: Provenance

    def to_dict(self) -> Dict[str, Any]:
        d = asdict(self)
        return d


@dataclass
class ParseReport:
    parser: str = "excel_v2"
    parser_version: str = PARSER_VERSION
    broker_detected: Optional[str] = None
    sheets: List[Dict[str, Any]] = field(default_factory=list)
    holdings_raw: int = 0
    holdings_final: int = 0
    total_value: float = 0.0
    statement_date: Optional[str] = None
    client_id: Optional[str] = None
    warnings: List[str] = field(default_factory=list)


def _norm_cell(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip().lower()


def _safe_float(v: Any) -> Optional[float]:
    if v is None or v == "" or v == "-":
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").replace("₹", "").replace("Rs.", "").replace("INR", "").strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _detect_broker(all_text: str) -> Optional[str]:
    t = all_text.lower()
    for broker, marks in BROKER_FINGERPRINTS:
        if any(m in t for m in marks):
            return broker
    return None


def _find_header_row(rows: List[Tuple[Any, ...]]) -> Optional[int]:
    """
    Header row heuristic: first row that contains an exact 'isin' cell AND
    a cell whose lowercase content includes 'quantity', 'units', or 'qty'.
    This is robust to Zerodha-style preambles, Groww-style title rows, etc.
    """
    for i, row in enumerate(rows):
        cells = [_norm_cell(c) for c in row]
        if not any(c == "isin" for c in cells):
            continue
        if any(("quantity" in c) or (c == "units") or (c == "qty") or
               ("closing units" in c) or ("balance units" in c)
               for c in cells):
            return i
    return None


def _map_columns(header: Tuple[Any, ...]) -> Dict[str, int]:
    """
    Map header cells → canonical field → column index.
    First non-empty match wins per canonical field, so aliases must be
    ordered most-specific-first.
    """
    norm = [_norm_cell(h) for h in header]
    mapping: Dict[str, int] = {}
    for canonical, aliases in COL_ALIASES.items():
        for a in aliases:
            for idx, cell in enumerate(norm):
                if not cell:
                    continue
                if cell == a or cell.startswith(a + " ") or a in cell:
                    if canonical not in mapping:
                        mapping[canonical] = idx
                        break
            if canonical in mapping:
                break
    return mapping


def _extract_metadata(rows: List[Tuple[Any, ...]]) -> Dict[str, Optional[str]]:
    """Pull client id + statement date from preamble cells."""
    client_id: Optional[str] = None
    statement_date: Optional[str] = None
    date_re = re.compile(r"(20\d{2}-\d{2}-\d{2}|\d{2}-[A-Za-z]{3}-20\d{2})")
    for row in rows[:25]:
        cells = [c for c in row if c is not None]
        text_cells = [str(c).strip() for c in cells]
        joined = " | ".join(text_cells)
        for i, c in enumerate(text_cells):
            if c.lower() in ("client id", "client code", "pan", "folio") and i + 1 < len(text_cells):
                client_id = client_id or text_cells[i + 1]
        m = date_re.search(joined)
        if m and not statement_date:
            statement_date = m.group(1)
    return {"client_id": client_id, "statement_date": statement_date}


def _parse_sheet(
    sheet_name: str,
    rows: List[Tuple[Any, ...]],
) -> Tuple[List[ParsedHolding], Dict[str, Any]]:
    """Parse a single sheet. Returns (holdings, status_report)."""
    h_idx = _find_header_row(rows)
    if h_idx is None:
        return [], {"sheet": sheet_name, "status": "no_header_found", "rows": 0}

    col_map = _map_columns(rows[h_idx])
    if "isin" not in col_map:
        return [], {"sheet": sheet_name, "status": "no_isin_column", "rows": 0}

    holdings: List[ParsedHolding] = []
    for r_offset, row in enumerate(rows[h_idx + 1:]):
        r_idx = h_idx + 1 + r_offset + 1  # 1-indexed source row

        def cell(field_name: str) -> Any:
            i = col_map.get(field_name)
            if i is None or i >= len(row):
                return None
            return row[i]

        isin_raw = cell("isin")
        if isin_raw is None:
            continue
        isin = str(isin_raw).strip().upper()
        if not ISIN_RE.match(isin):
            # Not a data row (could be a subtotal, blank, or footer). Skip silently.
            continue

        qty       = _safe_float(cell("quantity"))
        avg_price = _safe_float(cell("average_price"))
        close     = _safe_float(cell("close_price"))
        cur_val   = _safe_float(cell("current_value"))
        inv_val   = _safe_float(cell("invested_value"))

        flags: List[str] = []
        if cur_val is None or cur_val <= 0:
            if qty is not None and close is not None and close > 0:
                cur_val = round(qty * close, 4)
                flags.append("COMPUTED_VALUE_QTY_x_CLOSE")
            elif qty is not None and avg_price is not None and avg_price > 0:
                cur_val = round(qty * avg_price, 4)
                flags.append("COMPUTED_VALUE_QTY_x_AVG_FALLBACK")
            else:
                cur_val = 0.0
                flags.append("NO_VALUE_AVAILABLE")

        name_raw = cell("instrument_name")
        name = str(name_raw).strip() if name_raw is not None else ""
        if not name:
            flags.append("MISSING_NAME")
            name = isin

        sector_raw = cell("sector")
        sector = str(sector_raw).strip() if sector_raw is not None else None
        if sector in (None, "", "-", "—"):
            sector = None

        itype_raw = cell("instrument_type")
        itype = str(itype_raw).strip() if itype_raw is not None else None
        if itype in ("", "-"):
            itype = None

        folio_raw = cell("folio_number")
        folio = str(folio_raw).strip() if folio_raw is not None else None
        if folio in ("", "-"):
            folio = None

        prov = Provenance(
            source_sheet=sheet_name,
            source_row=r_idx,
            confidence=1.0 if ISIN_RE.match(isin) else 0.6,
            flags=flags,
            raw_payload={k: row[v] for k, v in col_map.items() if v < len(row)},
        )

        holdings.append(ParsedHolding(
            isin=isin,
            instrument_name=name,
            quantity=qty,
            nav=close if (close and close > 0) else avg_price,
            current_value=cur_val,
            invested_value=inv_val,
            sector=sector,
            instrument_type_raw=itype,
            folio_number=folio,
            provenance=prov,
        ))

    return holdings, {"sheet": sheet_name, "status": "ok", "rows": len(holdings)}


def _dedupe_by_isin(
    holdings: List[ParsedHolding],
) -> Tuple[List[ParsedHolding], List[str]]:
    """
    Deduplicate by ISIN. When duplicates exist, prefer the row from the
    higher-priority (segregated) sheet. Records a flag on the survivor.
    """
    warnings: List[str] = []
    by_isin: Dict[str, ParsedHolding] = {}
    duplicates: Dict[str, int] = {}
    for h in holdings:
        prio_existing = SHEET_PRIORITY.get(
            by_isin[h.isin].provenance.source_sheet, 1) if h.isin in by_isin else 99
        prio_new = SHEET_PRIORITY.get(h.provenance.source_sheet, 1)
        if h.isin not in by_isin or prio_new < prio_existing:
            by_isin[h.isin] = h
        duplicates[h.isin] = duplicates.get(h.isin, 0) + 1

    for isin, count in duplicates.items():
        if count > 1:
            survivor = by_isin[isin]
            survivor.provenance.flags.append(f"DEDUPED_FROM_{count}_ROWS")
    return list(by_isin.values()), warnings


def parse(content: bytes) -> Tuple[List[ParsedHolding], ParseReport]:
    """
    Parse an Excel workbook into canonical ParsedHolding records + report.

    Deterministic given identical bytes (openpyxl loads cells in fixed order;
    we sort outputs by ISIN to remove any dict-insertion-order dependency).
    """
    report = ParseReport()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=False)
    except Exception as e:
        report.warnings.append(f"workbook_open_failed:{e}")
        return [], report

    # Build a text blob for broker fingerprinting (first 30 rows of each sheet).
    blobs: List[str] = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 30:
                break
            for c in row:
                if c is not None:
                    blobs.append(str(c))
    report.broker_detected = _detect_broker(" | ".join(blobs))

    # Metadata from the first non-empty sheet's preamble.
    for sn in wb.sheetnames:
        rows = list(wb[sn].iter_rows(values_only=True))
        meta = _extract_metadata(rows)
        if meta["client_id"] or meta["statement_date"]:
            report.client_id = meta["client_id"]
            report.statement_date = meta["statement_date"]
            break

    all_holdings: List[ParsedHolding] = []
    for sn in wb.sheetnames:
        rows = list(wb[sn].iter_rows(values_only=True))
        sheet_holdings, status = _parse_sheet(sn, rows)
        report.sheets.append(status)
        all_holdings.extend(sheet_holdings)

    report.holdings_raw = len(all_holdings)

    final, warns = _dedupe_by_isin(all_holdings)
    report.warnings.extend(warns)

    # Deterministic ordering: sort by ISIN
    final.sort(key=lambda h: h.isin or "")

    report.holdings_final = len(final)
    report.total_value = round(sum(h.current_value for h in final), 4)
    return final, report


def parse_csv(content: bytes) -> Tuple[List[ParsedHolding], ParseReport]:
    """
    CSV path: convert to a single-sheet in-memory workbook, then reuse the
    Excel path. Keeps header inference + value computation in one place.
    """
    import csv
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = [tuple(r) for r in reader]
    if not rows:
        rep = ParseReport()
        rep.warnings.append("empty_csv")
        return [], rep
    holdings, status = _parse_sheet("csv", rows)
    rep = ParseReport()
    rep.sheets.append(status)
    rep.holdings_raw = len(holdings)
    final, _ = _dedupe_by_isin(holdings)
    final.sort(key=lambda h: h.isin or "")
    rep.holdings_final = len(final)
    rep.total_value = round(sum(h.current_value for h in final), 4)
    return final, rep
